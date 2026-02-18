from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import List, Optional
import re

from openpyxl import load_workbook


@dataclass
class PricingRule:
    match_type: str
    pattern: str
    utilidad_percent: Optional[Decimal]


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in value.lower() if ch.isalnum())


def load_rules(path: Path) -> List[PricingRule]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header = [cell.value or "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {}
    for idx, name in enumerate(header):
        key = _normalize_header(str(name))
        if key in {"matchtype", "tipo"}:
            header_map["match_type"] = idx
        elif key in {"pattern", "patron"}:
            header_map["pattern"] = idx
        elif key in {"utilidadpercent", "utilidad", "markuppercent", "markup"}:
            header_map["utilidad_percent"] = idx

    rules: List[PricingRule] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        match_type = str(row[header_map.get("match_type", 0)] or "contains").strip().lower()
        pattern = str(row[header_map.get("pattern", 1)] or "").strip()
        utilidad_raw = row[header_map.get("utilidad_percent", 2)] if "utilidad_percent" in header_map else None
        utilidad = Decimal(str(utilidad_raw)) if utilidad_raw not in (None, "") else None
        if not pattern:
            continue
        rules.append(PricingRule(match_type=match_type, pattern=pattern, utilidad_percent=utilidad))
    return rules


def find_rule(description: str, rules: List[PricingRule]) -> Optional[PricingRule]:
    if not description:
        return None
    text = description.lower()
    for rule in rules:
        pattern = rule.pattern.lower()
        if rule.match_type == "exact" and text == pattern:
            return rule
        if rule.match_type == "startswith" and text.startswith(pattern):
            return rule
        if rule.match_type == "regex":
            try:
                if re.search(rule.pattern, description, re.IGNORECASE):
                    return rule
            except re.error:
                continue
        if rule.match_type in ("contains", "contiene") and pattern in text:
            return rule
    return None
