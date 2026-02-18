from decimal import Decimal
from typing import Iterable, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .invoice_parser import InvoiceHeader
from .pricing import MarkupConfig, PriceRow


def _excel_round_to_step(expression: str, step: Decimal, mode: str) -> str:
    if step <= 0:
        return expression
    step_str = str(step)
    if mode == "nearest":
        rounded = f"ROUND(({expression})/{step_str},0)*{step_str}"
    elif mode == "down":
        rounded = f"FLOOR({expression},{step_str})"
    else:
        rounded = f"CEILING({expression},{step_str})"
    return f"IF(MOD({rounded},1000)=0,{rounded}+{step_str},{rounded})"


def _add_header_sheet(wb: Workbook, header: InvoiceHeader) -> None:
    ws = wb.create_sheet("Encabezado")
    ws.append(["Campo", "Valor"])
    rows = [
        ("Proveedor", header.supplier_name),
        ("NIT Proveedor", header.supplier_id),
        ("Cliente", header.customer_name),
        ("Factura", header.invoice_id),
        ("CUFE", header.cufe),
        ("Fecha factura", header.issue_date),
        ("Fecha vencimiento", header.due_date),
        ("Moneda", header.currency),
        ("Subtotal", header.subtotal),
        ("Impuestos", header.tax_total),
        ("Total con impuestos", header.total_tax_inclusive),
        ("Total factura", header.total),
    ]
    for label, value in rows:
        ws.append([label, value])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    money_fmt = "#,##0.00"
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell.value, (int, float, Decimal)):
            cell.number_format = money_fmt


def export_price_rows(
    rows: Iterable[PriceRow],
    output_path: str,
    sheet_name: str = "Productos",
    header: Optional[InvoiceHeader] = None,
    config: Optional[MarkupConfig] = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    config = config or MarkupConfig()

    headers = [
        "Linea factura",
        "Producto",
        "Cantidad",
        "IVA %",
        "Costo bruto unitario",
        "Costo neto unitario",
        "Venta bruta unitario",
        "Venta neto unitario",
        "Valor total Neto compra",
        "Descuento %",
    ]
    ws.append(headers)
    money_fmt = "#,##0.00"
    percent_fmt = "0.00"

    for idx, row in enumerate(rows, start=2):
        excel_row = [
            row.source_line_id,
            row.product,
            float(row.quantity),
            float(row.tax_percent),
            row.cost_bruto_unit,
            None,  # costo neto formula
            None,  # venta bruta formula
            None,  # venta neta formula
            None,  # total neto compra formula
            float(row.discount_percent),
        ]
        ws.append(excel_row)
        qty_cell = f"C{idx}"
        iva_cell = f"D{idx}"
        bruto_cell = f"E{idx}"
        net_cell = f"F{idx}"
        venta_bruta_cell = f"G{idx}"
        venta_neta_cell = f"H{idx}"
        total_neto_cell = f"I{idx}"
        desc_cell = f"J{idx}"

        ws.cell(row=idx, column=6).value = f"=ROUND({bruto_cell}*(1-{desc_cell}/100)*(1+{iva_cell}/100),2)"

        if row.markup_percent is not None:
            net_raw = f"{net_cell}*(1+{row.markup_percent}/100)"
        else:
            net_raw = f"IF({net_cell}<{config.threshold},{net_cell}/{config.below_divisor},{net_cell}*{config.above_multiplier})"
        net_rounded = _excel_round_to_step(net_raw, config.round_net_step, config.rounding_mode)
        ws.cell(row=idx, column=8).value = f"={net_rounded}"

        ws.cell(row=idx, column=7).value = f"={venta_neta_cell}/(1+{iva_cell}/100)"
        ws.cell(row=idx, column=9).value = f"=ROUND({net_cell}*{qty_cell},2)"

    for col_idx, col_header in enumerate(headers, start=1):
        column = get_column_letter(col_idx)
        ws.column_dimensions[column].width = max(len(col_header) + 2, 18)
    for row_cells in ws.iter_rows(min_row=2, min_col=3, max_col=10):
        # Cantidad
        row_cells[0].number_format = "0.00"
        # IVA %
        row_cells[1].number_format = percent_fmt
        # Costo/Venta columnas + total
        for cell in row_cells[2:7]:
            cell.number_format = money_fmt
        # Descuento %
        row_cells[7].number_format = percent_fmt

    if header:
        _add_header_sheet(wb, header)

    wb.save(str(output_path))
